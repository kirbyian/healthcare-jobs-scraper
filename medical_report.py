from bs4 import BeautifulSoup
import requests
import csv
import json
import email_report
import mater_parser
import sys
!{sys.executable} -m pip install openpyxl
from openpyxl import load_workbook

medical_search_list=['cardio','genetics','pharma','internal','derma','gastro',
                        'endocrinol','diabetes','general internal'
                        'genitourinary','geriatric','infectious disease','oncolog',
                        'nephrologo','neuro','pallative','rehab',
                         'resp','rheumatolog','gynaco','haem']

medical_substrings_to_depts={
    'cardio':'Cardiology',
    'genetics':'Clinical Genetics',
    'derma':'Dermatology',
    'endocrinol':'Endocrinology & Diabetes Mellitus',
    'diabetes':'Endocrinology & Diabetes Mellitus',
    'gynaco':'Obs & Gyna',
    'gastro':'Gastro & General Physician',
    'geriatric':'Geriatrics',
    'infectious disease':'Infectious Diseases',
    'oncolog':'Oncology',
    'resp':'Resp & GIM',
    'neuro':'Neurology',
    'rheumatolog':'Rheumatology & General Physician',
    'haem':'Haematology',
    'rehab':'Rehabilitation Medicine'
}

class BaseHTMLParser:
    
    def __init__(self):
        self.job_title = ""
        self.location = ""
        self.hospital = ""
        self.deadline = ""
        self.contact = ""
        self.duration = ""
        self.department=""
        self.match_found=False;

    def parse_job_page(self, job_url):
        raise NotImplementedError("Subclasses must implement the parse_job_page method.")

    def get_job_details(self, job_url):
        self.parse_job_page(job_url)
        return self.job_title, self.hospital, self.location,self.deadline, self.contact, self.duration,self.department,self.match_found

    def job_title_in_search_list(self,job_title):
        for substring in medical_search_list:
            if substring.lower() in job_title.lower():
                mapped_value = medical_substrings_to_depts.get(substring.lower(), "unknown")
                return mapped_value,True
        return '',False



class hse_parser(BaseHTMLParser):
    
    def scrape_job_data(self,base_url, pagination_base_url, writer, is_consultant):
        last_page_number = get_last_page_number(pagination_base_url)

        # Create an instance of the parser class
        parser_instance = hse_parser()

        for page_number in range(1, last_page_number + 1):
            page_url = f'{pagination_base_url}?pageNumber={page_number}'
            response = requests.get(page_url)
            soup = BeautifulSoup(response.content, 'html.parser')

            # Extract links to individual job pages using BeautifulSoup
            job_links = [base_url + link['href'].strip() for link in soup.select('p.read-more a.btn')]

            for job_link in job_links:
                job_title, hospital, location, deadline, contact, duration, department, match_found = parser_instance.get_job_details(job_link)
                if is_consultant:
                    job_title = 'Consultant'
                else:
                    job_title = 'Registrar'
                if match_found:
                    writer.writerow({'Position': job_title, 'Hospital': hospital, 'Location': location, 'Link for Post': job_link,
                                     'Deadline': deadline, 'Contact': contact, 'Duration': duration, 'Department': department})
    
    def parse_job_page(self, job_url):
            response = requests.get(job_url)
            soup = BeautifulSoup(response.text, 'html.parser')

            # Extract job title
            title_element = soup.find('td', {'headers': 'col2'})
            if title_element:
                self.job_title = title_element.get_text(strip=True)

            # Find the next td element after the job location for county
            location_element = soup.find('td', {'headers': 'col1', 'abbr': 'County'})
            next_location_element = location_element.find_next('td', {'headers': 'col2'})
            if next_location_element:
                self.location = next_location_element.get_text(strip=True)
    
            # Find the next td element after the job title for location
            hospital_element = soup.find('td', {'headers': 'col1', 'abbr': 'Location'})
            next_hospital_element = hospital_element.find_next('td')
            if next_hospital_element:
                self.hospital = next_hospital_element.get_text(strip=True)
    
            # Find the next td element after the job title for deadline
            deadline_element = soup.find('td', {'headers': 'col1', 'abbr': 'ClosingDate'})
            next_deadline_element = deadline_element.find_next('td')
            if next_deadline_element:
                self.deadline = next_deadline_element.get_text(strip=True)
    
            # Find the next td element after the job title for location
            contact_element = soup.find('td', {'headers': 'col1', 'abbr': 'Enquiries'})
            next_contact_element = contact_element.find_next('td')
            if next_contact_element:
                self.contact = next_contact_element.get_text(strip=True)
    
            # Find the next td element after the job title for duration
            duration_element = soup.find('td', {'headers': 'col1', 'abbr': 'Type'})
            next_duration_element = duration_element.find_next('td')
            if next_duration_element:
                duration_text = next_duration_element.get_text(strip=True).lower()
                self.duration = "Perm" if "perm" in duration_text else "Temp"
    
            # Todo: Determine Dept based on Position text
            self.department,self.match_found=self.job_title_in_search_list(self.job_title)

def get_last_page_number(base_url):
    response = requests.get(base_url)
    soup = BeautifulSoup(response.content, 'html.parser')
    last_page_element = soup.find('li', {'class': 'go-last'})
    if last_page_element:
        last_page_url = last_page_element.find('a')['href']
        last_page_number = int(last_page_url.split('=')[-1])
        return last_page_number
    return 1

def main():
    hse_consultants_url = 'https://www.hse.ie/eng/staff/jobs/job-search/medical-dental/consultants/'
    hse_non_consultants_url='https://www.hse.ie/eng/staff/jobs/job-search/medical-dental/nchd/sho-registrar/'
    hse_base_url = 'https://www.hse.ie'

    wb = load_workbook(filename = 'job_data.xlsx')
    
    # Select the active sheet (you can also specify a sheet by name if there are multiple sheets)
    sheet = wb.active

    # Specify the starting row where you want to write the data
    start_row = 3  # To append data after the existing content
    
    with open('job_data.csv', 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['Posted', 'Hospital', 'Position', 'Department', 'Location', 'Duration', 'Deadline', 'Contact', 'Link for Post']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        # HSE POSITIONS
        hse = hse_parser()
        hse.scrape_job_data(hse_base_url, hse_consultants_url, writer, True)
        hse.scrape_job_data(hse_base_url, hse_non_consultants_url, writer, False)

        # MATER POSITION
        mater_parser.scrape_job_data_mater(writer)
        

    # Read CSV data from the CSV file
    with open('job_data.csv', 'r', newline='') as csv_file:
        csv_reader = csv.reader(csv_file)

        # Skip the header row
        next(csv_reader)
        # Iterate through each row in the CSV and write it to the Excel sheet
        for csv_row in csv_reader:
            for col_num, value in enumerate(csv_row, start=1):
                sheet.cell(row=start_row, column=col_num, value=value)
            start_row += 1

    wb.save('job_data.xlsx')

    # Send email after writing is done
    with open('job_data.xlsx', 'rb') as f:
        file_content = f.read()
        email_report.send_email_with_attachment(file_content, 'job_data.xlsx')

if __name__ == "__main__":
    main()